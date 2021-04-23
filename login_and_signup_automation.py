from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
driver =  webdriver.Chrome(executable_path = "./chromedriver.exe")
driver.maximize_window()
driver.get("http://pepble.com/#/signIn")

# click on signup
driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/div[2]/div/a[1]").click()
time.sleep(2)
# load the excel sheet active page
wb_obj = openpyxl.load_workbook("./test_details.xlsx")
sheet_obj = wb_obj.active

for i in range(3,15):
    name = sheet_obj.cell(row = i, column = 1).value
    #print(name)
    mobile = sheet_obj.cell(row = i, column = 2).value
    #print(mobile)
    email = sheet_obj.cell(row = i, column = 3).value
    #print(email)
    pwd = sheet_obj.cell(row = i, column = 4).value
    #print(pwd)
    p_pwd = sheet_obj.cell(row = i, column = 5).value
    #print(p_pwd)
    # click on Name and get the entry from excel to test
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[1]/div[1]/input")
    time.sleep(1)
    element.send_keys(name)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[1]/div[2]/input")
    element.send_keys(mobile)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[2]/div/input")
    element.send_keys(email)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[3]/div/input")
    element.send_keys(pwd)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[4]/div/input")
    element.send_keys(p_pwd)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form/fieldset/div[5]/button")
    element.click()
    time.sleep(3)
    driver.get("http://pepble.com/#/signUp")

driver.get("http://pepble.com/#/signIn")
for i in range(18,30):
    user_name = sheet_obj.cell(row = i, column = 1).value
    #print(user_name)
    password = sheet_obj.cell(row = i, column = 2).value
    #print(password)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form[1]/fieldset/div[1]/div/input")
    element.send_keys(user_name)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form[1]/fieldset/div[2]/div/input")
    element.send_keys(password)
    time.sleep(2)
    element = driver.find_element_by_xpath("//*[@id='app']/div/main/section/div/div[2]/form[1]/fieldset/div[4]/button")
    element.click()
    driver.get("http://pepble.com/#/signIn")


#driver.implicitly_wait(10)

#driver.forward()
#driver.back()
#time.sleep(5)
driver.close()
